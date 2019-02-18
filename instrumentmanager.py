import asyncio
import sys
import time
import openpyxl
import visa

from os import listdir
from os.path import isfile, join

from PyQt5.QtCore import QObject

from analyzer import Analyzer
from analyzermock import AnalyzerMock
from generator import Generator
from generatormock import GeneratorMock
from source import Source
from sourcemock import SourceMock

# MOCK
mock_enabled = False

class InstrumentManager(QObject):

    excel_extension = '.xlsx'

    measure_freq_unit = 'GHz'
    measure_pow_unit = 'dBm'
    measure_current = 10

    control_voltage = 3.0   # was 4.75

    measure_freq = {
        1: {'f1': 0.01, 'f2': 0.02, 'f3': 0.03},
        2: {'f1': 0.01, 'f2': 0.02, 'f3': 0.03}
    }
    measure_pow = {
        1: {'p1': -40.0, 'p2': -30.0, 'p3': -20.0},
        2: {'p1': -40.0, 'p2': -30.0, 'p3': -20.0}
    }

    def __init__(self, parent=None, measureModels: dict=None):
        super(InstrumentManager, self).__init__(parent)
        print('instrument manager: init')

        self._source: SourceMock = None
        self._generator1: GeneratorMock = None
        self._generator2: GeneratorMock = None
        self._analyzer: AnalyzerMock = None

        self._samplePresent = False
        self._fileName = ''
        self._params = dict()

        self.pow_limit = -10
        print('lossess level:', self.pow_limit)

        self._generatorList = ['E4438C', 'N5181B', 'N5183A']

        self.measureModels = measureModels

    def findInstruments(self):
        print('instrument manager: find instruments')
        print('searching GPIB...')

        def find_live():
            # TODO refactor this mess
            rm = visa.ResourceManager()
            for res in rm.list_resources():
                try:
                    print(f'trying: {res}')
                    inst = rm.open_resource(res)
                    answer = inst.query('*IDN?')
                    model = answer.split(',')[1].strip()
                    print(model, self._generatorList)
                    if model == 'E3648A':
                    # if model == 'E3631A':
                        self._source = Source(res, answer, inst)
                        print('>>> source')
                    # elif 'N5183A' in answer:
                    #     self._generator1 = Generator(res, answer, inst)
                    elif model in self._generatorList:
                        if not self._generator1:
                            self._generator1 = Generator(res, answer, inst)
                            print('>>> gen1')
                        else:
                            self._generator2 = Generator(res, answer, inst)
                            print('>>> gen2')
                    elif model == 'N9030A':
                        self._analyzer = Analyzer(res, answer, inst)
                        print('>>> analyzer')
                except Exception as ex:
                    print(ex)

        def find_mocks():
            self._source = SourceMock('source addr', ',source mock,')
            self._generator1 = GeneratorMock('gen1 addr', ',gen1 mock,')
            self._generator2 = GeneratorMock('gen2 addr', ',gen2 mock,')
            self._analyzer = AnalyzerMock('analyzer addr', ',analyzer mock,')

        if mock_enabled:
            find_mocks()
        else:
            find_live()

        return self._source is not None and \
               self._generator1 is not None and \
               self._generator2 is not None and \
               self._analyzer is not None

    def getInstrumentNames(self):
        return self._source.name, self._generator1.name, self._generator2.name, self._analyzer.name

    def checkSample(self):
        print('instrument manager: check sample')

        if isfile('settings.ini'):
            with open('settings.ini') as f:
                line = f.readline()
                self.pow_limit = float(line.split('=')[1].strip())

        self._source.set_current(chan=1, value=40, unit='mA')
        self._source.set_voltage(chan=1, value=3, unit='V')
        self._source.set_current(chan=2, value=40, unit='mA')
        self._source.set_voltage(chan=2, value=0, unit='V')

        self._source.set_output(chan=1, state='ON')
        self._source.set_output(chan=2, state='ON')

        self._generator1.set_modulation('OFF')
        self._generator1.set_freq(100, 'MHz')
        self._generator1.set_pow(-20, 'dBm')
        self._generator1.set_output('ON')

        self._generator2.set_modulation('OFF')
        self._generator2.set_freq(90, 'MHz')
        self._generator2.set_pow(-10, 'dBm')
        self._generator2.set_output('ON')

        if not mock_enabled:
            time.sleep(1)

        self._analyzer.set_autocalibrate('OFF')
        self._analyzer.set_span(1, 'MHz')
        self._analyzer.set_measure_center_freq(100, 'MHz')
        self._analyzer.set_marker_mode(1, 'POS')
        self._analyzer.set_marker1_x_center(10, 'MHz')

        if not mock_enabled:
            time.sleep(1)

        read_pow = self._analyzer.read_pow(marker=1)
        print('marker value:', read_pow)

        if read_pow > self.pow_limit:
            self._samplePresent = True
            print('sample test ok')
        else:
            self._samplePresent = False
            print('losses too big, check sample connection')
        # self._samplePresent = True

        self._analyzer.set_autocalibrate('ON')
        self._generator1.set_output('OFF')
        self._source.set_output(1, 'OFF')
        self._source.set_output(2, 'OFF')

        if not mock_enabled:
            time.sleep(0.3)

        self._source.set_system_local()
        self._generator1.set_system_local()
        self._generator2.set_system_local()
        self._analyzer.set_system_local()

        return self._samplePresent

    def checkTaskTable(self):

        def getFileList(data_path):
            return [l for l in listdir(data_path) if isfile(join(data_path, l)) and self.excel_extension in l]

        # TODO: extract measurement manager class
        print('instrument manager: check task table')

        files = getFileList('.')
        length = len(files)
        if length > 1:
            print('too many task tables, abort')
            return False
        elif length <= 0:
            print('no task table found, abort')
            return False

        self._fileName = files[0]
        print('using task table:', self._fileName)

        wb = openpyxl.load_workbook(self._fileName)
        ws = wb.active

        cols = ws.max_column
        rows = ws.max_row
        letters = [chr(ch_code).upper() for ch_code in range(ord('a'), ord('z') + 1)]
        start_col = 2

        count = int((rows - 1) / 3)

        # TODO: validate table
        for row in range(count):
            self._params[row + 1] = [[ws[char + str(row * 3 + 1 + num)].value for num in range(1, 4)] for char in letters[start_col:cols]]

        headers = ['№'] + [ws[char + '1'].value for char in letters[start_col:cols]]

        for model in self.measureModels.values():
            model.initHeader(headers)

        wb.close()
        print('done read')
        return True

    def measure(self, letter):
        print(f'start measure letter={letter}')

        self.measureTask(letter)

        self.measureModels[letter].initModel(self._params[letter])

        print('end measure')

    def generator_freq_sweep(self, generator, letter):

        def do_branch(self, letter, branch):

            if branch == 1:
                chan1voltage = self.control_voltage
                chan2voltage = 0.0
            else:
                chan1voltage = 0.0
                chan2voltage = self.control_voltage

            self._source.set_voltage(1, chan1voltage, 'V')
            self._source.set_voltage(2, chan2voltage, 'V')
            self._source.set_output(1, 'ON')
            self._source.set_output(2, 'ON')

            generator.set_output('ON')

            pows = list()

            for freq in self.measure_freq[letter].values():
                generator.set_freq(freq, self.measure_freq_unit)
                self._analyzer.set_measure_center_freq(freq, self.measure_freq_unit)
                self._analyzer.set_marker1_x_center(freq, self.measure_freq_unit)

                pows.append(self._analyzer.read_pow(1))
                if not mock_enabled:
                    time.sleep(0.3)

            generator.set_output('OFF')
            self._source.set_output(1, 'OFF')
            self._source.set_output(2, 'OFF')

            return pows

        generator.set_pow(self.measure_pow[1]['p1'], self.measure_pow_unit)

        [do_branch(self, letter, branch) for branch in [1, 2]]
        [do_branch(self, letter, branch) for branch in [1, 2]]
        return [do_branch(self, letter, branch) for branch in [1, 2]]

    def measureTask(self, letter: int):

        print(f'measurement task run, letter={letter}')

        self._generator1.set_modulation('OFF')
        self._generator2.set_modulation('OFF')

        self._analyzer.set_autocalibrate('OFF')
        self._analyzer.set_span(10, 'MHz')

        self._source.set_current(1, 40, 'mA')
        self._source.set_current(2, 40, 'mA')

        self._analyzer.set_marker_mode(1, 'POS')

        # gen 1 freq sweep
        pows = self.generator_freq_sweep(self._generator1, letter)
        print(pows)

        # gen 2 freq sweep
        pows = self.generator_freq_sweep(self._generator2, letter)
        print(pows)

        # gen 1 pow sweep
        freq = self.measure_freq[letter]['f3']
        freq_unit = self.measure_freq_unit
        pow_unit = self.measure_pow_unit

        self._generator1.set_freq(freq, freq_unit)

        self._analyzer.set_autocalibrate('OFF')
        self._analyzer.set_span(1, 'MHz')
        self._analyzer.set_marker_mode(1, 'POS')
        self._analyzer.set_measure_center_freq(freq, 'MHz')
        self._analyzer.set_marker1_x_center(freq, 'MHz')

        self._source.set_voltage(1, 3, 'V')
        self._source.set_voltage(2, 0, 'V')
        self._source.set_output(1, 'ON')
        self._source.set_output(2, 'ON')

        self._generator1.set_output('ON')

        for pow in self.measure_pow[letter].values():
            self._generator1.set_pow(pow, pow_unit)

            if not mock_enabled:
                time.sleep(0.3)

            read_pow = self._analyzer.read_pow(1)
            print(read_pow)

        self._generator1.set_output('OFF')

        self._source.set_output(1, 'OFF')
        self._source.set_output(2, 'OFF')
        self._source.set_system_local()

        self._analyzer.set_autocalibrate('ON')
